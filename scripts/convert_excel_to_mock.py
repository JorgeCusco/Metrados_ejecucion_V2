#!/usr/bin/env python3
import openpyxl
import json
from uuid import uuid4

def process_sheet(ws, sheet_name):
    """Procesa una hoja del Excel y retorna array de partidas en formato mock"""
    partidas = []
    codigo_to_id = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        # Extraer solo primeras 4 columnas: Código, Descripción, Unidad, Modificación
        codigo = row[0]  # Columna A
        descripcion = row[1]  # Columna B
        unidad = row[2]  # Columna C
        modificacion = row[3]  # Columna D
        
        # Saltar filas vacías o s in código
        if not codigo or str(codigo).strip() == '':
            continue
        
        codigo_str = str(codigo).strip()
        descripcion_str = str(descripcion).strip() if descripcion else ''
        unidad_str = str(unidad).strip() if unidad else ''
        modificacion_str = str(modificacion).strip() if modificacion else 'SM'
        
        # Validación: los códigos deben tener puntos (ej: OE.1)
        if '.' not in codigo_str:
            continue
        
        # Generar ID único
        record_id = str(uuid4())
        codigo_to_id[codigo_str] = record_id
        
        # Determinar nivel jerárquico por puntos en codigo
        nivel = len(codigo_str.split('.'))
        
        # Determinar si es título (si la unidad está vacía)
        es_titulo = unidad_str == ''
        
        # Construir jerarquía (array de textos con antecedentes)
        jerarquia = []
        
        partida = {
            "id": record_id,
            "codigo": codigo_str,
            "descripcion": descripcion_str,
            "unidad": unidad_str,
            "es_titulo": es_titulo,
            "parent_id": None,  # Se resuelve después
            "nivel_jerarquia": nivel,
            "jerarquia": jerarquia,
            "especialidad": "hospital" if sheet_name.lower().startswith('hosp') else "contingencia",
            "modificacion": modificacion_str,
            "apu": None
        }
        
        partidas.append(partida)
    
    # Segunda pasada: resolver parent_id y jerarquía
    for partida in partidas:
        codigo = partida['codigo']
        puntos = codigo.split('.')
        
        if len(puntos) > 1:
            # Buscar padre quitando el último nivel
            padre_codigo = '.'.join(puntos[:-1])
            if padre_codigo in codigo_to_id:
                partida['parent_id'] = codigo_to_id[padre_codigo]
        
        # Construir jerarquía buscando ancestros
        jerarquia = []
        for i in range(1, len(puntos)):
            ancestor_codigo = '.'.join(puntos[:i])
            for p in partidas:
                if p['codigo'] == ancestor_codigo:
                    jerarquia.append(f"{p['codigo']} {p['descripcion']}")
                    break
        partida['jerarquia'] = jerarquia
    
    return partidas

# Procesar Excel
print("Leyendo nueva_data.xlsx...")
wb = openpyxl.load_workbook('../nueva_data.xlsx')
print(f"Hojas disponibles: {wb.sheetnames}")

# Encontrar las hojas correctas
hospital_sheet = None
contingencia_sheet = None

for sheet_name in wb.sheetnames:
    if sheet_name.upper().startswith('HOSP'):
        hospital_sheet = sheet_name
    elif sheet_name.upper().startswith('CONTING'):
        contingencia_sheet = sheet_name

if not hospital_sheet or not contingencia_sheet:
    print(f"ERROR: No se encontraron hojas Hospital y Contingencia")
    print(f"Hojas encontradas: {wb.sheetnames}")
    exit(1)

print(f"Procesando: {hospital_sheet} y {contingencia_sheet}")

hospital_partidas = process_sheet(wb[hospital_sheet], hospital_sheet)
contingencia_partidas = process_sheet(wb[contingencia_sheet], contingencia_sheet)

print(f"\n✅ {hospital_sheet}: {len(hospital_partidas)} partidas")
print(f"✅ {contingencia_sheet}: {len(contingencia_partidas)} partidas")

# Mostrar ejemplos
print(f"\n=== Ejemplos {hospital_sheet} (primeros 3) ===")
for p in hospital_partidas[:3]:
    print(f"  {p['codigo']}: {p['descripcion'][:40]}... [{p['unidad']}] mod={p['modificacion']}")

print(f"\n=== Ejemplos {contingencia_sheet} (primeros 3) ===")
for p in contingencia_partidas[:3]:
    print(f"  {p['codigo']}: {p['descripcion'][:40]}... [{p['unidad']}] mod={p['modificacion']}")

# Guardar para próximo paso
with open('hospital_converted.json', 'w', encoding='utf-8') as f:
    json.dump(hospital_partidas, f, ensure_ascii=False, indent=2)

with open('contingencia_converted.json', 'w', encoding='utf-8') as f:
    json.dump(contingencia_partidas, f, ensure_ascii=False, indent=2)

print("\n✅ Datos guardados en hospital_converted.json y contingencia_converted.json")
