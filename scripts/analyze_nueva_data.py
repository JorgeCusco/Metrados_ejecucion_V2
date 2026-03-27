#!/usr/bin/env python3
import openpyxl
import json

wb = openpyxl.load_workbook('../nueva_data.xlsx')
print('Hojas disponibles:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== {sheet_name} ===')
    print(f'Filas: {ws.max_row}, Columnas: {ws.max_column}')
    print('\nPrimeras 10 filas:')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        print(f'Fila {i}: {row}')
